import io
import pandas as pd
import re

from reportlab.platypus import (
    PageBreak,
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import inch

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


def convert_df_to_excel(df):

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name="Report"
        )

        workbook = writer.book
        worksheet = writer.sheets["Report"]

        # ---------------------------------
        # Header Style
        # ---------------------------------

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
            size=12
        )

        # ---------------------------------
        # Cell Style
        # ---------------------------------

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        wrap_alignment = Alignment(
            wrap_text=True,
            vertical="top",
            horizontal="left"
        )

        # ---------------------------------
        # Header Formatting
        # ---------------------------------

        for cell in worksheet[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = wrap_alignment

        # ---------------------------------
        # Data Formatting
        # ---------------------------------

        for row in worksheet.iter_rows(min_row=2):

            for cell in row:

                cell.border = thin_border
                cell.alignment = wrap_alignment

        # ---------------------------------
        # Auto Column Width
        # ---------------------------------

        for column_cells in worksheet.columns:

            max_length = 0

            column = column_cells[0].column

            for cell in column_cells:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except:
                    pass

            adjusted_width = min(max_length + 5, 50)

            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = adjusted_width

        # ---------------------------------
        # Freeze Header
        # ---------------------------------

        worksheet.freeze_panes = "A2"

        # ---------------------------------
        # Auto Filter
        # ---------------------------------

        worksheet.auto_filter.ref = worksheet.dimensions

        # ---------------------------------
        # Row Height
        # ---------------------------------

        for row in worksheet.iter_rows():

            worksheet.row_dimensions[
                row[0].row
            ].height = 35

        # ---------------------------------
        # Status Coloring
        # ---------------------------------

        status_column = None

        for cell in worksheet[1]:

            if cell.value == "Status":

                status_column = cell.column

                break

        if status_column:

            green = PatternFill(
                fill_type="solid",
                start_color="C6EFCE",
                end_color="C6EFCE"
            )

            yellow = PatternFill(
                fill_type="solid",
                start_color="FFF2CC",
                end_color="FFF2CC"
            )

            red = PatternFill(
                fill_type="solid",
                start_color="F4CCCC",
                end_color="F4CCCC"
            )

            for row in range(2, worksheet.max_row + 1):

                cell = worksheet.cell(row=row, column=status_column)

                value = str(cell.value).strip().title()

                if value == "Covered":

                    cell.fill = green

                elif value == "Partial":

                    cell.fill = yellow

                elif value == "Missing":

                    cell.fill = red

    return output.getvalue()



def convert_rtm_to_pdf(df, coverage, covered, partial, missing):

    output = io.BytesIO()

    doc = SimpleDocTemplate(
        output,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    title_style = styles["Heading1"]
    title_style.alignment = TA_CENTER

    heading_style = styles["Heading2"]

    normal_style = styles["BodyText"]

    story = []

    # ======================================
    # TITLE
    # ======================================

    story.append(Paragraph(
        "AI QA Assistant",
        title_style
    ))

    story.append(
        Paragraph(
            "Requirement Traceability Report",
            heading_style
        )
    )

    story.append(Spacer(1, 0.35 * inch))

    # ======================================
    # DASHBOARD
    # ======================================

    dashboard_data = [

        ["Metric", "Value"],

        ["Coverage", f"{coverage}%"],

        ["Covered", covered],

        ["Partial", partial],

        ["Missing", missing]

    ]

    dashboard = Table(
        dashboard_data,
        colWidths=[220, 120]
    )

    dashboard.setStyle(

        TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.darkblue),

            ("TEXTCOLOR", (0,0), (-1,0), colors.white),

            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

            ("FONTSIZE", (0,0), (-1,-1), 11),

            ("ALIGN", (0,0), (-1,-1), "CENTER"),

            ("GRID", (0,0), (-1,-1), 1, colors.black),

            ("BOTTOMPADDING", (0,0), (-1,0), 10),

            ("BACKGROUND", (0,1), (-1,-1), colors.beige)

        ])

    )

    story.append(dashboard)

    story.append(Spacer(1, 0.4 * inch))

    # ======================================
    # OVERALL VERDICT
    # ======================================

    if coverage >= 90:

        verdict = "🟢 Excellent Requirement Coverage"

    elif coverage >= 70:

        verdict = "🟡 Requirement Needs Minor Improvements"

    else:

        verdict = "🔴 High Risk Requirement"

    story.append(

        Paragraph(

            f"<b>Overall Verdict:</b> {verdict}",

            heading_style

        )

    )

    story.append(Spacer(1, 0.3 * inch))

    # ======================================
    # REQUIREMENTS
    # ======================================

    for index, row in df.iterrows():

        story.append(

            Paragraph(

                f"<b>Requirement {index+1}</b>",

                heading_style

            )

        )

        story.append(

            Paragraph(

                f"<b>Requirement</b><br/>{row['Requirement']}",

                normal_style

            )

        )

        story.append(Spacer(1, 0.1 * inch))

        story.append(

            Paragraph(

                f"<b>Status:</b> {row['Status']}",

                normal_style

            )

        )

        story.append(Spacer(1, 0.1 * inch))

        story.append(

            Paragraph(

                "<b>Missing Scenarios</b>",

                normal_style

            )

        )

        missing = format_ai_text(row["Missing Scenario"])

        story.append(
            Paragraph(
                missing,
                normal_style
            )
        )

        story.append(Spacer(1, 0.1 * inch))

        story.append(

            Paragraph(

                "<b>Recommendations</b>",

                normal_style

            )

        )

        recommendation = format_ai_text(row["Recommendation"])

        story.append(
            Paragraph(
                recommendation,
                normal_style
            )
        )

        story.append(Spacer(1, 0.35 * inch))

    # ======================================
    # SUMMARY
    # ======================================

    story.append(PageBreak())

    story.append(

        Paragraph(

            "AI Summary",

            title_style

        )

    )

    story.append(Spacer(1, 0.25 * inch))

    if coverage >= 90:

        summary = """
        The generated test cases provide excellent coverage of the supplied
        requirements. The application appears ready for QA execution with only
        minor improvements recommended.
        """

    elif coverage >= 70:

        summary = """
        The requirements have good overall coverage. However, additional
        negative, boundary, and security scenarios should be considered to
        improve testing completeness.
        """

    else:

        summary = """
        Significant gaps were identified in the supplied requirements.
        Additional requirement clarification and comprehensive test case
        generation are recommended before test execution.
        """

    story.append(

        Paragraph(

            summary,

            normal_style

        )

    )

    story.append(Spacer(1, 0.25 * inch))

    story.append(

        Paragraph(

            "<b>Generated by AI QA Assistant using Gemini AI</b>",

            heading_style

        )

    )

    doc.build(story)

    output.seek(0)

    return output.getvalue()



def format_ai_text(text):

    text = str(text)

    # Put numbered points on new lines
    text = re.sub(r'(\d+\.)', r'<br/><br/><b>\1</b>', text)

    # Put bullets on new lines
    text = text.replace("•", "<br/>•")

    return text.strip()